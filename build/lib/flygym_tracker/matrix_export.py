"""One workbook per run: a sheet per parameter, vials down, timepoints across.

WHAT THE RIG OWNER ASKED FOR, and why it is a different shape from the CSVs rather than a
replacement for them.

    "one big excel table where there will be a tab for each parameter, and inside a tab the data
     will be a table where the rows are vials (1:32 labeled as number and A or B) and the columns
     are the consecutive acquired timepoints"

THE CSVs ARE LONG; THIS IS WIDE. `activity.csv` and `behaviour.csv` are one row per vial per bin --
the right shape for a machine to append to during a three-day run, and the right shape for pandas
afterwards, but the wrong shape for a person opening a file to see what happened. A 32-row sheet
whose columns march forward in time is something you can read across, sort, and paste into a plot
without a groupby. Both are kept: the CSVs are still written as the run goes, and this is built
from them at the end. If the workbook write fails, no measurement is lost.

WHY IT IS BUILT AT THE END RATHER THAN APPENDED. A sheet like this grows a COLUMN per timepoint,
and neither openpyxl nor Excel is happy being asked to add a column to a 30 000-column sheet every
ten seconds for three days. The CSVs take the streaming load; this is the readable snapshot.

EVERY VIAL GETS A ROW, INCLUDING ONES THAT NEVER REPORTED. A missing vial silently absent from the
sheet reads as "there were 30 vials"; an empty row reads as "this vial reported nothing", which is
the fact. Cells with no measurement are left EMPTY rather than zeroed, for the reason that runs
through this whole project: an empty vial has no height, which is not the same claim as "its flies
are at the bottom".
"""
from __future__ import annotations

import math
import os
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

#: Vial rows, in order: A1..A16 then B1..B16. The LABEL carries the face, so a row is identifiable
#: on its own once the sheet is pasted somewhere else.
VIALS_PER_FACE = 16
FACES = ("A", "B")


def vial_labels() -> List[str]:
    return ["%s%d" % (face, i + 1) for face in FACES for i in range(VIALS_PER_FACE)]


def vial_label(face: str, vial_id: int) -> str:
    """`A1`..`A16`, `B1`..`B16` from a face and a 1-based GLOBAL vial id.

    Global ids are ``face_index * 16 + local`` with local running 1..16, so face B's first vial is
    17. The -1 is what makes vial 16 the sixteenth of its face rather than the first of the next --
    the same off-by-one that put face A's vial 16 in the cell labelled B1 on the live grid.
    """
    local = (int(vial_id) - 1) % VIALS_PER_FACE + 1
    return "%s%d" % (face, local)


def _is_number(value) -> bool:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return not math.isnan(number)


def run_stamp(started_iso: Optional[str] = None) -> str:
    """`YYYYmmdd-HHMMSS` for a filename, from the run's START.

    THE START, NOT THE END, and not "now". Every file of one run then carries the same stamp, so a
    directory of several runs sorts and groups by eye; naming them at write time would scatter one
    run's files across the timeline and give the workbook a different stamp from the CSVs.
    """
    if started_iso:
        try:
            return datetime.fromisoformat(started_iso).strftime("%Y%m%d-%H%M%S")
        except (TypeError, ValueError):
            pass
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def build_matrix(rows: Iterable[dict], field: str, *, decimals: int = 4
                 ) -> Tuple[List[float], Dict[str, Dict[float, float]]]:
    """`(timepoints, {vial_label: {t: value}})` for one parameter.

    Timepoints are the distinct `elapsed_s` values that carry a measurement for this field, sorted.
    Rounding them is what makes activity's bin times and behaviour's dwell times line up as
    columns instead of drifting apart by microseconds.
    """
    times: set = set()
    table: Dict[str, Dict[float, float]] = {label: {} for label in vial_labels()}
    for row in rows or ():
        value = row.get(field)
        if not _is_number(value):
            continue
        try:
            elapsed = round(float(row["elapsed_s"]), decimals)
            label = vial_label(str(row["face"]), int(row["vial_id"]))
        except (KeyError, TypeError, ValueError):
            continue
        if label not in table:
            continue
        times.add(elapsed)
        # LAST WRITE WINS for a repeated (vial, time): behaviour rows are per dwell and activity
        # rows per bin, so a collision means two measurements of the same field at the same
        # timestamp, which only happens if a caller mixes sources. Averaging them silently would
        # invent a third number.
        table[label][elapsed] = float(value)
    return sorted(times), table


def write_workbook(path, sheets: Dict[str, Sequence[dict]], *, meta: Optional[dict] = None,
                   fields: Optional[Sequence[Tuple[str, str]]] = None) -> str:
    """Write one workbook: a sheet per parameter, plus a `run` sheet of the metadata.

    Args:
        path: the .xlsx to write.
        sheets: ``{source_name: rows}`` -- e.g. the activity rows and the behaviour rows. Every
            field in `fields` is looked for across all of them, so a parameter that only one
            source produces still gets its own sheet.
        meta: run metadata, written verbatim to a `run` sheet. KEPT, as asked: the timestamps and
            the config snapshot are what make a workbook self-describing months later.
        fields: ``[(field, label)]`` to export. Defaults to everything plottable.

    Returns the path written.
    """
    from openpyxl import Workbook

    if fields is None:
        from flygym_tracker.gui.behaviour_series import PLOTTABLE

        fields = PLOTTABLE

    combined: List[dict] = []
    for rows in (sheets or {}).values():
        combined.extend(rows or ())

    book = Workbook()
    book.remove(book.active)                      # drop the default empty sheet

    # THE METADATA SHEET IS FIRST, so a workbook opens on what this run WAS rather than on a wall
    # of numbers with no units, no start time and no config behind them.
    info = book.create_sheet("run")
    info.append(["key", "value"])
    for key, value in _flatten(meta or {}):
        info.append([key, value])

    written = 0
    for field, label in fields:
        times, table = build_matrix(combined, field)
        if not times:
            # NO SHEET FOR A PARAMETER NOTHING MEASURED. An empty grid of 32 rows and no columns
            # invites the reading "every vial was zero here", which is the opposite of the truth.
            continue
        sheet = book.create_sheet(_sheet_name(field))
        sheet.append(["vial"] + [round(t, 3) for t in times])
        for vial in vial_labels():
            values = table.get(vial, {})
            # EMPTY, NOT ZERO, for a timepoint this vial did not report -- see the module docstring.
            sheet.append([vial] + [values.get(t) for t in times])
        sheet.freeze_panes = "B2"                 # keep the vial column and the time row in view
        sheet["A1"].comment = None
        written += 1
    if not written:
        book.create_sheet("no data")

    directory = os.path.dirname(str(path))
    if directory:
        os.makedirs(directory, exist_ok=True)
    book.save(str(path))
    return str(path)


def _sheet_name(field: str) -> str:
    """Excel sheet names are capped at 31 chars and cannot contain ``[]:*?/\\``."""
    name = str(field)
    for bad in "[]:*?/\\":
        name = name.replace(bad, "_")
    return name[:31] or "sheet"


def _flatten(data, prefix: str = "") -> List[Tuple[str, object]]:
    """`{a: {b: 1}}` -> ``[("a.b", 1)]``, so a nested config snapshot fits a two-column sheet."""
    out: List[Tuple[str, object]] = []
    if isinstance(data, dict):
        for key, value in data.items():
            out.extend(_flatten(value, "%s.%s" % (prefix, key) if prefix else str(key)))
    elif isinstance(data, (list, tuple)):
        out.append((prefix, ", ".join(str(v) for v in data)))
    else:
        out.append((prefix, data))
    return out
