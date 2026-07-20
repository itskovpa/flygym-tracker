"""The workbook: a sheet per parameter, vials down, consecutive timepoints across.

    "one big excel table where there will be a tab for each parameter, and inside a tab the data
     will be a table where the rows are vials (1:32 labeled as number and A or B) and the columns
     are the consecutive acquired timepoints"

THE LOAD-BEARING TEST IS THE LAST ONE: it runs a REAL pipeline and asserts every cell of the
workbook against the CSV the run wrote. This is where the claim that used to belong to the live
table -- what you are shown is what was recorded -- now lives.
"""
from __future__ import annotations

import csv
import glob
import os

import numpy as np
import pytest

from flygym_tracker.matrix_export import (build_matrix, run_stamp, vial_label, vial_labels,
                                          write_workbook)


def _row(elapsed, face, vial_id, **fields):
    row = {"run_id": "r", "elapsed_s": elapsed, "face": face, "vial_id": vial_id}
    row.update(fields)
    return row


# =============================================================================================
# Labels and shape
# =============================================================================================
def test_the_rows_are_the_thirty_two_vials_labelled_by_face():
    labels = vial_labels()
    assert len(labels) == 32
    assert labels[0] == "A1" and labels[15] == "A16"
    assert labels[16] == "B1" and labels[31] == "B16"


def test_global_ids_map_onto_their_own_face():
    """Global ids are `face_index*16 + local` with local 1..16 -- 1-BASED. Face B's first vial is
    17. The same off-by-one that put face A's vial 16 in the cell labelled B1 on the live grid
    would put it in the wrong ROW here, silently, for every parameter."""
    assert vial_label("A", 1) == "A1"
    assert vial_label("A", 16) == "A16"
    assert vial_label("B", 17) == "B1"
    assert vial_label("B", 32) == "B16"


def test_every_vial_gets_a_row_even_if_it_never_reported(tmp_path):
    """A vial silently absent reads as "there were 30 vials"; an empty row reads as "this vial
    reported nothing", which is the fact."""
    from openpyxl import load_workbook

    path = tmp_path / "w.xlsx"
    write_workbook(path, {"a": [_row(0.0, "A", 1, motion_px_sum=5)]},
                   fields=[("motion_px_sum", "motion")])
    sheet = load_workbook(path)["motion_px_sum"]
    assert sheet.max_row == 33, "expected a header plus 32 vial rows"
    assert [sheet.cell(r, 1).value for r in range(2, 34)] == vial_labels()


def test_a_missing_measurement_is_empty_not_zero(tmp_path):
    """The rule that runs through this whole project: an empty vial has no height, which is not
    the same claim as "its flies are at the bottom"."""
    from openpyxl import load_workbook

    path = tmp_path / "w.xlsx"
    write_workbook(path, {"a": [_row(0.0, "A", 1, motion_px_sum=5)]},
                   fields=[("motion_px_sum", "motion")])
    sheet = load_workbook(path)["motion_px_sum"]
    assert sheet.cell(2, 2).value == 5           # A1 reported
    assert sheet.cell(3, 2).value is None        # A2 did not


def test_columns_are_the_consecutive_timepoints():
    times, table = build_matrix(
        [_row(0.0, "A", 1, x=1), _row(10.0, "A", 1, x=2), _row(5.0, "A", 2, x=3)], "x")
    assert times == [0.0, 5.0, 10.0], "timepoints are not sorted and de-duplicated"
    assert table["A1"] == {0.0: 1.0, 10.0: 2.0}
    assert table["A2"] == {5.0: 3.0}


def test_a_parameter_nothing_measured_gets_no_sheet(tmp_path):
    """An empty grid of 32 rows and no columns invites the reading "every vial was zero here",
    which is the opposite of the truth."""
    from openpyxl import load_workbook

    path = tmp_path / "w.xlsx"
    write_workbook(path, {"a": [_row(0.0, "A", 1, motion_px_sum=5)]},
                   fields=[("motion_px_sum", "motion"), ("never_measured", "nope")])
    assert "never_measured" not in load_workbook(path).sheetnames


def test_nan_is_not_a_measurement():
    _times, table = build_matrix([_row(0.0, "A", 1, x=float("nan")),
                                  _row(1.0, "A", 1, x=4.0)], "x")
    assert table["A1"] == {1.0: 4.0}


# =============================================================================================
# Metadata and naming
# =============================================================================================
def test_the_metadata_sheet_comes_first_and_keeps_the_run_details(tmp_path):
    """Kept, as asked. A workbook opening on a wall of numbers with no start time, no units and
    no config behind them is not self-describing three months later."""
    from openpyxl import load_workbook

    path = tmp_path / "w.xlsx"
    write_workbook(path, {"a": [_row(0.0, "A", 1, motion_px_sum=1)]},
                   meta={"run_id": "abc", "start_iso": "2026-01-01T09:00:00",
                         "meta": {"config": {"activity": {"pixel_threshold": 12.0}}}},
                   fields=[("motion_px_sum", "motion")])
    book = load_workbook(path)
    assert book.sheetnames[0] == "run"
    keys = {book["run"].cell(r, 1).value: book["run"].cell(r, 2).value
            for r in range(2, book["run"].max_row + 1)}
    assert keys["run_id"] == "abc"
    assert keys["start_iso"] == "2026-01-01T09:00:00"
    assert keys["meta.config.activity.pixel_threshold"] == 12.0, "nested config was not flattened"


def test_the_stamp_comes_from_the_run_start():
    """THE START, not the end and not "now": every file of one run then carries the same stamp, so
    a directory of several runs sorts and groups by eye."""
    assert run_stamp("2026-07-20T14:05:09") == "20260720-140509"
    assert len(run_stamp(None)) == len("20260720-140509")
    assert len(run_stamp("not a date")) == len("20260720-140509")


def test_every_output_file_carries_the_run_stamp(tmp_path):
    from flygym_tracker.logger import ActivityLogger

    logger = ActivityLogger(output_dir=str(tmp_path), run_id="r", fmt="csv", rolling="daily")
    for path in (logger.activity_path(), logger.events_path(), logger.behaviour_path(),
                 logger.meta_path(), logger.workbook_path()):
        assert logger.stamp in path.name, "%s has no run stamp" % path.name


# =============================================================================================
# The claim that matters: the workbook IS the recorded data
# =============================================================================================
def test_the_workbook_matches_the_csv_the_run_wrote(tmp_path):
    """END TO END through a real `TrackerPipeline`. A workbook showing plausible-but-different
    numbers would be worse than none: it is the file a person actually opens."""
    from openpyxl import load_workbook

    from flygym_tracker.calibration import (build_two_face_calibration_from_polygons,
                                            load_calibration, save_calibration)
    from flygym_tracker.config import load_config
    from flygym_tracker.logger import ActivityLogger
    from flygym_tracker.pipeline import TrackerPipeline
    from flygym_tracker.types import Frame

    H, W = 200, 300
    polygons = [[[20 + 90 * i, 40], [90 + 90 * i, 40],
                 [90 + 90 * i, 160], [20 + 90 * i, 160]] for i in range(3)]

    def frame(i):
        image = np.full((H, W), 200, dtype=np.uint8)
        image[150 - 2 * (i % 30):158 - 2 * (i % 30), 45:55] = 40
        return image

    frames = [frame(i) for i in range(40)]

    class Source:
        fps = 20.0

        def __init__(self):
            self.i = 0

        def open(self):
            pass

        def close(self):
            pass

        def read(self):
            if self.i >= len(frames):
                return None
            out = Frame(image=frames[self.i], index=self.i, t_monotonic=self.i / 20.0,
                        t_wall_iso="2026-01-01T00:00:%06.3f" % (self.i / 20.0))
            self.i += 1
            return out

    calib, masks, _ = build_two_face_calibration_from_polygons(
        polygons, frames[0], (W, H), faces=("A", "B"))
    bundle = str(tmp_path / "calib")
    save_calibration(calib, masks, bundle)
    out = str(tmp_path / "out")
    logger = ActivityLogger(output_dir=out, run_id="t", fmt="csv", rolling="daily", meta={})
    config = load_config(overrides={"binning": {"bin_seconds": 0.5},
                                    "activity": {"pixel_threshold": 1.0},
                                    "rotation": {"detector": "adaptive"}})
    pipeline = TrackerPipeline(config=config, calibration=load_calibration(bundle),
                               source=Source(), logger=logger, marker_detector=None,
                               clock="index")
    summary = pipeline.run()

    assert summary["workbook"], "no workbook was written: %s" % summary["workbook_error"]
    assert not summary["workbook_error"]
    book = load_workbook(summary["workbook"])
    assert "motion_px_sum" in book.sheetnames

    rows = []
    for path in sorted(glob.glob(os.path.join(out, "activity_*.csv"))):
        with open(path, newline="", encoding="utf-8") as f:
            rows.extend(csv.DictReader(f))
    assert rows, "the run wrote no activity csv"

    sheet = book["motion_px_sum"]
    times = [sheet.cell(1, c).value for c in range(2, sheet.max_column + 1)]
    labels = [sheet.cell(r, 1).value for r in range(2, sheet.max_row + 1)]
    for record in rows:
        label = vial_label(record["face"], int(record["vial_id"]))
        column = times.index(round(float(record["elapsed_s"]), 3)) + 2
        row_index = labels.index(label) + 2
        assert sheet.cell(row_index, column).value == pytest.approx(
            float(record["motion_px_sum"])), (
            "%s at t=%s: workbook %r vs csv %r" % (label, record["elapsed_s"],
                                                   sheet.cell(row_index, column).value,
                                                   record["motion_px_sum"]))
